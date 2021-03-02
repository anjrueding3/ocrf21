import jpg2data
from openpyxl import load_workbook, Workbook


def updateExcel(objectList, targetExcel):
    workbook = load_workbook(targetExcel)
    sheet = workbook.active
    maxRow = sheet.max_row
    
    for object in objectList:
        matchFound = False

        #match object.PO() to Cell B(of row Num) until there's a match
        for i in range(maxRow):
            #start from A1, B1, C1
            actualRow = i + 1
            actualIHD = 'A' + str(actualRow)
            actualPO = 'B' + str(actualRow)
            actualStyle = 'C' + str(actualRow)
            
            if object.getPO() == sheet[actualPO].value:
                sheet[actualIHD] = object.getIHD()
                sheet[actualStyle] = object.getStyle()
                matchFound = True
                break
        #if there's no match, then append a new row of data. And increase maxRow variable +1.
        if matchFound == False:
            sheet.append([object.getIHD(), object.getPO(), object.getStyle()])
            maxRow += 1

    workbook.save(filename= targetExcel)




